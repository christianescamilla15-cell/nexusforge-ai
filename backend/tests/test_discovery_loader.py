"""Tests for Gap F — discovery_loader.py.

Covers:
- Parsing of fase-3-maestro.md format (H-NNN headers, states, domains, apps)
- Parsing of fase-2-por-dominio format (DOMAIN-NNN headers)
- Dedup: maestro findings take precedence over domain-local ones
- Classification: blockers, quick wins, corrections detected by keywords
- Statistics: totals, domains_covered, apps_mentioned
- Index queries: findings_for_app, findings_for_domain, blockers_by_app
- Graceful handling: missing directory, empty files, malformed headers
- Integration with the real committed corpus (if available)
"""
from __future__ import annotations

from pathlib import Path

import pytest

from app.refactor.discovery_loader import (
    DiscoveryFinding,
    DiscoveryIndex,
    load_discovery_context,
    load_entregables_context,
    _parse_maestro,
    _parse_domain_file,
    _parse_trazabilidad_xlsx,
    _app_codename_from_label,
    _normalize_severity,
    _OPENPYXL_AVAILABLE,
)


# ── Fixtures ──────────────────────────────────────────────────────────


_MAESTRO_SAMPLE = """\
## Hallazgos — Categoría: TEST

### H-001 — Pipeline end-to-end del ciclo del boleto
- **Estado**: `VALIDATED`
- **Dominios**: ventas, volado
- **Menciones**: 6
- **Descripción**: PNR→Boleto→cintas→contabilidad. Tier 0.
- **Relación NexusForge**: `tenant-alpha` modela parcialmente

### H-002 — NO existe orquestador centralizado
- **Estado**: `VALIDATED`
- **Dominios**: miatech, governance
- **Descripción**: "No tenemos algo así. Cobertura vía humanos." Bloqueador operativo.
- **Relación NexusForge**: **NEW-GAP** — orquestador como componente

### H-003 — Robot cierre vuelos downgraded (quick win)
- **Estado**: `VALIDATED`
- **Dominios**: ventas
- **Descripción**: Robot ya existe, solo está inhibido. Quick win de reactivación.

### H-004 — `app-05` Directas ≠ Indirectas (CORRIGE modelo)
- **Estado**: `CONFIRMED-BY-DOCS`
- **Dominios**: comisiones
- **Descripción**: P5 = Directas AS400 COBOL. Lo de Lucia = Indirectas AWS. NO es lo mismo.
- **Relación NexusForge**: refinar recipe `app-05`
"""


_DOMAIN_SAMPLE = """\
# Fase 2 — Dominio: ventas

## Hallazgos en este dominio

### VENTAS-001 — Canal directo vs indirecto
- **Estado**: `VALIDATED`
- **Descripción**: 50% corporate, agencias (AmEx, Carlson).

### VENTAS-002 — Fraude detectado $43M
- **Estado**: `RAW`
- **Descripción**: Nelly mencionó hallazgo de módulo. ROI claro para business case.
"""


# ── Parser: maestro ──────────────────────────────────────────────────


def test_parse_maestro_extracts_all_findings():
    findings = _parse_maestro(_MAESTRO_SAMPLE)
    assert len(findings) == 4
    ids = [f.id for f in findings]
    assert "H-001" in ids
    assert "H-002" in ids
    assert "H-003" in ids
    assert "H-004" in ids


def test_parse_maestro_states():
    findings = _parse_maestro(_MAESTRO_SAMPLE)
    by_id = {f.id: f for f in findings}
    assert by_id["H-001"].state == "VALIDATED"
    assert by_id["H-004"].state == "CONFIRMED-BY-DOCS"


def test_parse_maestro_domains():
    findings = _parse_maestro(_MAESTRO_SAMPLE)
    by_id = {f.id: f for f in findings}
    assert "ventas" in by_id["H-001"].domains
    assert "volado" in by_id["H-001"].domains
    assert "governance" in by_id["H-002"].domains


def test_parse_maestro_detects_blockers():
    findings = _parse_maestro(_MAESTRO_SAMPLE)
    by_id = {f.id: f for f in findings}
    assert by_id["H-002"].is_blocker is True
    assert by_id["H-001"].is_blocker is False


# ── Entregables xlsx ingestion ───────────────────────────────────────


def test_normalize_severity_maps_labels():
    assert _normalize_severity("CRITICAL") == "critical"
    assert _normalize_severity("blocker") == "critical"
    assert _normalize_severity("High") == "high"
    assert _normalize_severity("Major") == "high"
    assert _normalize_severity("MEDIUM") == "medium"
    assert _normalize_severity("info") == "low"
    assert _normalize_severity(None) == "medium"
    assert _normalize_severity("weird-label") == "medium"


def test_app_codename_explicit_takes_precedence():
    assert _app_codename_from_label("app-03-arc") == ["app-03-arc"]
    assert _app_codename_from_label("workshop ARC con app-03-bsp") == ["app-03-bsp"]


def test_app_codename_colloquial_mapping():
    assert "app-01" in _app_codename_from_label("SICOFAV")
    assert "app-03-arc" in _app_codename_from_label("ARC")
    assert "app-03-bsp" in _app_codename_from_label("BSP")
    assert "app-03-special" in _app_codename_from_label("Reembolsos Especiales")
    assert "app-04" in _app_codename_from_label("CFDIs")
    assert "app-05" in _app_codename_from_label("Comisiones Directas")
    assert "praxis-ecosystem" in _app_codename_from_label("Ecosistema PRAXIS")
    # AMBOS should return both ARC + BSP
    both = _app_codename_from_label("AMBOS")
    assert "app-03-arc" in both and "app-03-bsp" in both


@pytest.mark.skipif(not _OPENPYXL_AVAILABLE, reason="openpyxl not installed")
def test_parse_trazabilidad_xlsx_synthetic(tmp_path):
    """Build a synthetic trazabilidad workbook and verify parsing."""
    import openpyxl
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # ARC sheet
    ws = wb.create_sheet("Reembolsos ARC (P3)")
    ws.append(["Reembolsos ARC — Plataforma USA", None, None, None, None, None])
    ws.append(["#", "App", "Hallazgo", "Resumen / Sintesis", "Severidad", "Fuente"])
    ws.append([1, "ARC", "JWT hardcodeado (CWE-798)", "Secret key en codigo Python", "CRITICAL", "Informe ARC"])
    ws.append([2, "ARC", "CORS wildcard (CWE-942)", "Access-Control-Allow-Origin: *", "CRITICAL", "Informe ARC"])
    ws.append([3, "ARC", "Stack Python/Flask", "Info del stack", "INFO", "BIA"])

    # PRAXIS transversal sheet
    ws2 = wb.create_sheet("Ecosistema PRAXIS")
    ws2.append(["Hallazgos Ecosistema", None, None, None, None, None])
    ws2.append(["#", "App", "Hallazgo", "Resumen", "Severidad", "Fuente"])
    ws2.append([1, "PRAXIS", "189,997 issues totales", "Satellite + core COBOL", "CRITICAL", "Report Out Softtek"])

    path = tmp_path / "Hallazgos_Trazabilidad_test.xlsx"
    wb.save(path)

    findings = _parse_trazabilidad_xlsx(path)
    assert len(findings) == 4

    # ARC findings should have app-03-arc
    arc_jwt = next(f for f in findings if "JWT" in f.title)
    assert "app-03-arc" in arc_jwt.apps_affected
    assert arc_jwt.severity == "critical"
    assert arc_jwt.is_blocker is True
    assert arc_jwt.state == "CONFIRMED-BY-DOCS"

    # Low-severity INFO row is not a blocker
    stack_row = next(f for f in findings if "Stack" in f.title)
    assert stack_row.severity == "low"
    assert stack_row.is_blocker is False

    # PRAXIS ecosystem mapping
    praxis = next(f for f in findings if "189,997" in f.title)
    assert "praxis-ecosystem" in praxis.apps_affected


@pytest.mark.skipif(not _OPENPYXL_AVAILABLE, reason="openpyxl not installed")
def test_load_entregables_context_missing_dir(tmp_path):
    """Missing entregables directory returns empty list, not an error."""
    result = load_entregables_context(tmp_path / "does-not-exist")
    assert result == []


@pytest.mark.skipif(not _OPENPYXL_AVAILABLE, reason="openpyxl not installed")
def test_load_discovery_context_integrates_entregables(tmp_path):
    """Verify load_discovery_context picks up entregables/ xlsx files."""
    import openpyxl
    analisis = tmp_path / "analisis"
    analisis.mkdir()
    (analisis / "fase-3-maestro.md").write_text(_MAESTRO_SAMPLE, encoding="utf-8")

    entregables = analisis / "entregables-13abr"
    entregables.mkdir()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("SICOFAV (P1)")
    ws.append(["SICOFAV — header", None, None, None, None, None])
    ws.append(["#", "App", "Hallazgo", "Resumen", "Severidad", "Fuente"])
    ws.append([1, "SICOFAV", "Credenciales en web.config", "CWE-522", "CRITICAL", "FindingsExamples"])
    wb.save(entregables / "Hallazgos_Trazabilidad_5_Apps.xlsx")

    idx = load_discovery_context(tmp_path)
    # Should have 4 from maestro + 1 from entregables
    assert idx.total_findings == 5
    trz_ids = [f.id for f in idx.findings if f.id.startswith("TRZ")]
    assert len(trz_ids) == 1
    sicofav_f = next(f for f in idx.findings if f.id.startswith("TRZ"))
    assert sicofav_f.severity == "critical"
    assert "app-01" in sicofav_f.apps_affected


def test_parse_maestro_detects_quick_wins():
    findings = _parse_maestro(_MAESTRO_SAMPLE)
    by_id = {f.id: f for f in findings}
    assert by_id["H-003"].is_quick_win is True
    assert by_id["H-001"].is_quick_win is False


def test_parse_maestro_detects_corrections():
    findings = _parse_maestro(_MAESTRO_SAMPLE)
    by_id = {f.id: f for f in findings}
    assert by_id["H-004"].is_correction is True
    assert by_id["H-001"].is_correction is False


def test_parse_maestro_description():
    findings = _parse_maestro(_MAESTRO_SAMPLE)
    by_id = {f.id: f for f in findings}
    assert "PNR" in by_id["H-001"].description


def test_parse_maestro_nexusforge_relation():
    findings = _parse_maestro(_MAESTRO_SAMPLE)
    by_id = {f.id: f for f in findings}
    assert "tenant-alpha" in by_id["H-001"].nexusforge_relation


# ── Parser: domain file ──────────────────────────────────────────────


def test_parse_domain_file():
    findings = _parse_domain_file(_DOMAIN_SAMPLE, "ventas")
    assert len(findings) == 2
    assert findings[0].id == "VENTAS-001"
    assert findings[0].domains == ["ventas"]
    assert findings[1].id == "VENTAS-002"


def test_parse_domain_file_states():
    findings = _parse_domain_file(_DOMAIN_SAMPLE, "ventas")
    assert findings[0].state == "VALIDATED"
    assert findings[1].state == "RAW"


# ── Full loader ──────────────────────────────────────────────────────


def test_load_discovery_context_from_fixture(tmp_path: Path):
    """Build a mini analisis/ directory and verify the full loader."""
    analisis = tmp_path / "analisis"
    analisis.mkdir()
    (analisis / "fase-3-maestro.md").write_text(_MAESTRO_SAMPLE, encoding="utf-8")

    domain_dir = analisis / "fase-2-por-dominio"
    domain_dir.mkdir()
    (domain_dir / "ventas.md").write_text(_DOMAIN_SAMPLE, encoding="utf-8")
    (domain_dir / "_TEMPLATE.md").write_text("skip me", encoding="utf-8")

    index = load_discovery_context(tmp_path)

    # 4 from maestro + 2 from domain (not duplicated because they use different IDs)
    assert index.total_findings == 6
    assert index.total_validated >= 3
    assert index.total_blockers >= 1
    assert index.total_quick_wins >= 1
    assert index.total_corrections >= 1
    assert "ventas" in index.domains_covered


def test_load_discovery_context_dedup(tmp_path: Path):
    """Maestro finding H-001 should NOT be duplicated if it also
    appears in a domain file with the same ID."""
    analisis = tmp_path / "analisis"
    analisis.mkdir()
    maestro = "### H-001 — Test\n- **Estado**: `VALIDATED`\n- **Dominios**: ventas\n"
    (analisis / "fase-3-maestro.md").write_text(maestro, encoding="utf-8")

    domain_dir = analisis / "fase-2-por-dominio"
    domain_dir.mkdir()
    # Same ID in domain file — should be skipped
    domain = "### H-001 — Test duplicate\n- **Estado**: `RAW`\n"
    (domain_dir / "ventas.md").write_text(domain, encoding="utf-8")

    index = load_discovery_context(tmp_path)
    assert index.total_findings == 1
    assert index.findings[0].state == "VALIDATED"  # maestro version wins


def test_load_discovery_context_missing_dir(tmp_path: Path):
    """Missing analisis/ directory should return empty index, not error."""
    index = load_discovery_context(tmp_path)
    assert index.total_findings == 0
    assert index.findings == []


def test_load_discovery_context_empty_maestro(tmp_path: Path):
    """Empty maestro file should return empty index."""
    analisis = tmp_path / "analisis"
    analisis.mkdir()
    (analisis / "fase-3-maestro.md").write_text("# Empty\n", encoding="utf-8")

    index = load_discovery_context(tmp_path)
    assert index.total_findings == 0


# ── Index queries ────────────────────────────────────────────────────


def test_index_findings_for_app():
    index = DiscoveryIndex(findings=[
        DiscoveryFinding(id="H-1", apps_affected=["app-01", "app-02"]),
        DiscoveryFinding(id="H-2", apps_affected=["app-02"]),
        DiscoveryFinding(id="H-3", apps_affected=["app-03"]),
    ])
    assert len(index.findings_for_app("app-02")) == 2
    assert len(index.findings_for_app("app-99")) == 0


def test_index_findings_for_domain():
    index = DiscoveryIndex(findings=[
        DiscoveryFinding(id="H-1", domains=["ventas", "comisiones"]),
        DiscoveryFinding(id="H-2", domains=["ventas"]),
        DiscoveryFinding(id="H-3", domains=["governance"]),
    ])
    assert len(index.findings_for_domain("ventas")) == 2


def test_index_blockers_by_app():
    index = DiscoveryIndex(findings=[
        DiscoveryFinding(id="H-1", is_blocker=True, apps_affected=["app-01"]),
        DiscoveryFinding(id="H-2", is_blocker=True, apps_affected=["app-01", "app-02"]),
        DiscoveryFinding(id="H-3", is_blocker=False, apps_affected=["app-01"]),
    ])
    blockers = index.blockers_by_app
    assert len(blockers["app-01"]) == 2
    assert len(blockers["app-02"]) == 1


def test_index_to_json():
    index = DiscoveryIndex(
        findings=[DiscoveryFinding(id="H-1", title="test")],
        total_findings=1,
    )
    j = index.to_json()
    assert '"H-1"' in j
    assert '"test"' in j


# ── Integration with real corpus (if available) ─────────────────────


def test_real_corpus_if_available():
    """If the real NexusForge-Shared corpus is present on this machine,
    verify the loader can parse it end-to-end. Skip gracefully if not
    present (CI environment, fresh clone, etc.)."""
    real_path = Path(r"C:\Users\DANNY\NexusForge-Shared\transcripciones-generales")
    maestro = real_path / "analisis" / "fase-3-maestro.md"

    if not maestro.is_file():
        pytest.skip("Real corpus not available on this machine")

    index = load_discovery_context(real_path)

    # Sanity checks based on what we know from the 2026-04-12 pipeline run
    assert index.total_findings >= 80, f"Expected 80+ findings, got {index.total_findings}"
    assert index.total_validated >= 60
    assert len(index.domains_covered) >= 10
    assert "ventas" in index.domains_covered
    assert "miatech" in index.domains_covered
    assert index.total_blockers >= 1
    assert index.total_quick_wins >= 1
    assert index.total_corrections >= 1
