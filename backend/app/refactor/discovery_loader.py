"""Discovery context loader — Gap F.

Parses transcription analysis files (from the corpus pipeline in
``NexusForge-Shared/transcripciones-generales/analisis/``) and
produces a structured ``DiscoveryIndex`` that other NexusForge
modules can consume as prior knowledge.

The primary consumer is the ``StranglerPlanner`` (Gap 3) — when
discovery context is available, the planner can:

- Prioritize apps that have more blockers in the discovery corpus
- Honor decisions/corrections found in meetings (e.g., "app-05
  Directas ≠ Indirectas", "app-01 has a V2 in parallel")
- Weight risk using real meeting-sourced data (not just code metrics)
- Include quick wins identified verbally by domain owners

The loader is intentionally **read-only** — it never modifies the
analysis files. It also never reads the raw transcriptions (.txt /
.srt / .log) — only the structured Markdown outputs of the 5-phase
analysis pipeline.

Input: a directory path containing the ``analisis/`` subfolder with
phase-2 and/or phase-3 files.

Output: a ``DiscoveryIndex`` dataclass with:
- ``findings``: list of ``DiscoveryFinding`` (id, title, state,
  domains, apps_affected, severity, source_files)
- ``blockers_by_app``: dict mapping app codenames to their blockers
- ``quick_wins``: list of findings tagged as quick wins
- ``corrections``: list of findings that correct the synthetic model
- ``stats``: counts and metadata
"""
from __future__ import annotations

import json
import logging
import re
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

try:
    import openpyxl  # type: ignore
    _OPENPYXL_AVAILABLE = True
except ImportError:  # pragma: no cover
    _OPENPYXL_AVAILABLE = False


# ── Data model ────────────────────────────────────────────────────────


@dataclass
class DiscoveryFinding:
    """One structured finding extracted from the analysis pipeline."""

    id: str                                 # "H-001" or "VENTAS-003"
    title: str = ""
    state: str = "RAW"                      # RAW | VALIDATED | MAPPED | NEW-GAP | CONFIRMED-BY-DOCS | etc.
    domains: list[str] = field(default_factory=list)
    apps_affected: list[str] = field(default_factory=list)  # codenames like "app-01"
    severity: str = "medium"                # low | medium | high | critical
    description: str = ""
    source_files: list[str] = field(default_factory=list)  # relative paths to phase-1/2 files
    is_blocker: bool = False
    is_quick_win: bool = False
    is_correction: bool = False             # corrects something in the synthetic model
    nexusforge_relation: str = ""           # free text: what NexusForge component this maps to


@dataclass
class DiscoveryIndex:
    """Structured index of all discovery findings from the corpus pipeline."""

    findings: list[DiscoveryFinding] = field(default_factory=list)
    total_findings: int = 0
    total_validated: int = 0
    total_blockers: int = 0
    total_quick_wins: int = 0
    total_corrections: int = 0
    domains_covered: list[str] = field(default_factory=list)
    apps_mentioned: list[str] = field(default_factory=list)
    source_path: str = ""

    @property
    def blockers_by_app(self) -> dict[str, list[DiscoveryFinding]]:
        """Map app codenames to their blocker findings."""
        result: dict[str, list[DiscoveryFinding]] = {}
        for f in self.findings:
            if f.is_blocker:
                for app in f.apps_affected:
                    result.setdefault(app, []).append(f)
        return result

    @property
    def quick_wins(self) -> list[DiscoveryFinding]:
        return [f for f in self.findings if f.is_quick_win]

    @property
    def corrections(self) -> list[DiscoveryFinding]:
        return [f for f in self.findings if f.is_correction]

    def findings_for_app(self, codename: str) -> list[DiscoveryFinding]:
        """All findings that mention a specific app codename."""
        return [f for f in self.findings if codename in f.apps_affected]

    def findings_for_domain(self, domain: str) -> list[DiscoveryFinding]:
        """All findings that belong to a specific domain."""
        return [f for f in self.findings if domain in f.domains]

    def to_dict(self) -> dict[str, Any]:
        return {
            "total_findings": self.total_findings,
            "total_validated": self.total_validated,
            "total_blockers": self.total_blockers,
            "total_quick_wins": self.total_quick_wins,
            "total_corrections": self.total_corrections,
            "domains_covered": self.domains_covered,
            "apps_mentioned": self.apps_mentioned,
            "source_path": self.source_path,
            "findings": [asdict(f) for f in self.findings],
        }

    def to_json(self, indent: int = 2) -> str:
        return json.dumps(self.to_dict(), indent=indent, ensure_ascii=False)


# ── Parsing helpers ───────────────────────────────────────────────────

# Matches "### H-001 — <title>" or "### H-APP-01-V2 — <title>"
_FINDING_HEADER_RE = re.compile(
    r"^###\s+(H-[\w-]+)\s*[—–-]\s*(.+)$", re.MULTILINE
)

# Matches "- **Estado**: `VALIDATED`"
_STATE_RE = re.compile(r"\*\*Estado\*\*:\s*`([^`]+)`")

# Matches "- **Dominios**: ventas, comisiones"
_DOMAINS_RE = re.compile(r"\*\*Dominios?\*\*:\s*(.+)")

# Matches app codenames like "app-01", "app-03-special"
_APP_CODENAME_RE = re.compile(r"\bapp-\d{2}(?:-\w+)?\b")

# Quick win indicators
_QUICK_WIN_INDICATORS = [
    "quick win", "quick-win", "quickwin",
    "low-hanging", "ROI claro", "fácil de implementar",
    "ya existe.*solo.*inhibido", "robot.*downgraded",
]
_QUICK_WIN_RE = re.compile("|".join(_QUICK_WIN_INDICATORS), re.IGNORECASE)

# Blocker indicators
_BLOCKER_INDICATORS = [
    "bloqueador", "blocker", "bloqueo",
    "no existe.*dev", "no existe.*QA", "sin.*ambiente",
    "single point of failure", "no redundancia",
    "descontinuado", "discontinued",
]
_BLOCKER_RE = re.compile("|".join(_BLOCKER_INDICATORS), re.IGNORECASE)

# Correction indicators
_CORRECTION_INDICATORS = [
    "CORRIGE", "CORREGIDO", "CORRECTED", "NO es",
    "≠", "!=", "distinto", "confundió",
]
_CORRECTION_RE = re.compile("|".join(_CORRECTION_INDICATORS), re.IGNORECASE)


def _parse_maestro(text: str) -> list[DiscoveryFinding]:
    """Parse fase-3-maestro.md into a list of DiscoveryFinding."""
    findings: list[DiscoveryFinding] = []

    # Split the text into blocks per finding header
    headers = list(_FINDING_HEADER_RE.finditer(text))
    if not headers:
        return findings

    for i, match in enumerate(headers):
        finding_id = match.group(1).strip()
        title = match.group(2).strip()

        # Extract the block between this header and the next (or EOF)
        start = match.end()
        end = headers[i + 1].start() if i + 1 < len(headers) else len(text)
        block = text[start:end]

        # State
        state_match = _STATE_RE.search(block)
        state = state_match.group(1) if state_match else "RAW"

        # Domains
        domains_match = _DOMAINS_RE.search(block)
        domains: list[str] = []
        if domains_match:
            raw_domains = domains_match.group(1)
            domains = [d.strip().lower() for d in re.split(r"[,;]", raw_domains) if d.strip()]

        # Apps affected (scan the whole block for codenames)
        apps = sorted(set(_APP_CODENAME_RE.findall(block)))

        # Description (first line of the block that starts with "- **Descripción**:")
        desc_match = re.search(r"\*\*Descripci[oó]n\*\*:\s*(.+)", block)
        description = desc_match.group(1).strip() if desc_match else ""

        # NexusForge relation
        rel_match = re.search(r"\*\*Relaci[oó]n.*NexusForge\*\*:\s*(.+)", block)
        relation = rel_match.group(1).strip() if rel_match else ""

        # Classify: blocker, quick win, correction
        full_text = title + " " + block
        is_blocker = bool(_BLOCKER_RE.search(full_text))
        is_quick_win = bool(_QUICK_WIN_RE.search(full_text))
        is_correction = bool(_CORRECTION_RE.search(full_text))

        # Severity heuristic based on state + blocker + keywords
        severity = "medium"
        if is_blocker:
            severity = "high"
        if "critical" in block.lower() or "single point of failure" in block.lower():
            severity = "critical"
        if state in ("RAW",):
            severity = "low"

        findings.append(DiscoveryFinding(
            id=finding_id,
            title=title,
            state=state,
            domains=domains,
            apps_affected=apps,
            severity=severity,
            description=description[:500],
            is_blocker=is_blocker,
            is_quick_win=is_quick_win,
            is_correction=is_correction,
            nexusforge_relation=relation[:300],
        ))

    return findings


def _parse_domain_file(text: str, domain_name: str) -> list[DiscoveryFinding]:
    """Parse a fase-2-por-dominio/<domain>.md file.

    Falls back to a simpler parser than the maestro because domain
    files may not have the full H-NNN IDs yet — they use local IDs
    like VENTAS-001.
    """
    findings: list[DiscoveryFinding] = []

    # Match "### DOMAIN-NNN — <title>" pattern
    header_re = re.compile(
        r"^###\s+([\w-]+-\d{3})\s*[—–-]\s*(.+)$", re.MULTILINE
    )

    headers = list(header_re.finditer(text))
    for i, match in enumerate(headers):
        finding_id = match.group(1).strip()
        title = match.group(2).strip()

        start = match.end()
        end = headers[i + 1].start() if i + 1 < len(headers) else len(text)
        block = text[start:end]

        state_match = _STATE_RE.search(block)
        state = state_match.group(1) if state_match else "RAW"

        apps = sorted(set(_APP_CODENAME_RE.findall(block)))

        desc_match = re.search(r"\*\*Descripci[oó]n\*\*:\s*(.+)", block)
        description = desc_match.group(1).strip() if desc_match else ""

        full_text = title + " " + block
        is_blocker = bool(_BLOCKER_RE.search(full_text))
        is_quick_win = bool(_QUICK_WIN_RE.search(full_text))
        is_correction = bool(_CORRECTION_RE.search(full_text))

        findings.append(DiscoveryFinding(
            id=finding_id,
            title=title,
            state=state,
            domains=[domain_name],
            apps_affected=apps,
            severity="high" if is_blocker else "medium",
            description=description[:500],
            is_blocker=is_blocker,
            is_quick_win=is_quick_win,
            is_correction=is_correction,
        ))

    return findings


# ── Entregables ingestion (Excels from eng-partner deliverables) ───────

_SEVERITY_MAP = {
    "critical": "critical",
    "blocker": "critical",
    "high": "high",
    "major": "high",
    "medium": "medium",
    "med": "medium",
    "low": "low",
    "info": "low",
    "informational": "low",
}


def _normalize_severity(value: str | None) -> str:
    """Map spreadsheet severity labels to the loader's severity vocabulary."""
    if not value:
        return "medium"
    key = str(value).strip().lower()
    return _SEVERITY_MAP.get(key, "medium")


def _app_codename_from_label(label: str | None) -> list[str]:
    """Infer app codenames from a free-text label.

    Handles both explicit codenames ("app-03-arc") and colloquial labels
    that the engineering-partner spreadsheets use in rows.

    Naming convention: the IF-conditions below match incoming external
    label text (parser inputs from third-party Excel deliverables) and
    always emit our sanitized codenames. Real labels live here only as
    *input keys*; outputs are codenames. Per the F-02 audit fix, do not
    return any non-codename string from this function.
    """
    if not label:
        return []
    text = str(label).strip()
    explicit = sorted(set(_APP_CODENAME_RE.findall(text)))
    if explicit:
        return explicit

    # Colloquial → codename mapping (sourced from analisis/README.md).
    # The mapper accepts the real spreadsheet labels that `eng-partner`
    # uses as inputs and always emits our codenames. These inputs are
    # what the loader receives; the outputs are the sanitized codenames.
    upper = text.upper()
    colloquial: list[str] = []
    if "SICOFAV" in upper or "P1" in upper:
        colloquial.append("app-01")
    if "SRG" in upper or "P2" in upper:
        colloquial.append("app-02")
    if "ARC" in upper and "BSP" not in upper:
        colloquial.append("app-03-arc")
    if "BSP" in upper and "ARC" not in upper:
        colloquial.append("app-03-bsp")
    if "ESPECIAL" in upper:
        colloquial.append("app-03-special")
    if "AMBOS" in upper or upper == "ARC/BSP":
        colloquial.extend(["app-03-arc", "app-03-bsp"])
    if "CFDI" in upper or "P4" in upper:
        colloquial.append("app-04")
    if "COMISIONES" in upper or "P5" in upper:
        colloquial.append("app-05")
    if "PRAXIS" in upper or "ECOSISTEMA" in upper:
        colloquial.append("praxis-ecosystem")
    return colloquial


def _parse_trazabilidad_xlsx(
    path: Path,
    id_prefix: str = "TRZ",
) -> list[DiscoveryFinding]:
    """Parse an ``eng-partner`` Hallazgos_Trazabilidad-style workbook.

    Expected row format (per sheet):
        # | App | Hallazgo | Resumen / Sintesis | Severidad | Fuente / Documento

    Each sheet typically corresponds to one app (app-01, ARC, BSP, ...)
    plus an "Ecosistema PRAXIS" transversal sheet.
    """
    if not _OPENPYXL_AVAILABLE:
        logger.warning("openpyxl not installed — skipping xlsx ingestion at %s", path)
        return []

    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        logger.warning("Failed to open workbook %s: %s", path, exc)
        return []

    findings: list[DiscoveryFinding] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        domain_guess = _domain_from_sheet_name(sheet_name)
        header_seen = False
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if not row or all(c is None for c in row):
                continue
            # Look for header row: ['#', 'App', 'Hallazgo', 'Resumen...', 'Severidad', 'Fuente...']
            if not header_seen:
                cells = [str(c).strip().lower() if c else "" for c in row]
                if "hallazgo" in cells and ("severidad" in cells or "severity" in cells):
                    header_seen = True
                continue

            # Data row — expect 6 columns
            if len(row) < 5:
                continue
            num, app_label, hallazgo, resumen, severidad = row[:5]
            fuente = row[5] if len(row) > 5 else None

            if hallazgo is None:
                continue

            title = str(hallazgo).strip()
            if not title or title.lower().startswith("cobertura:"):
                continue

            # Build finding
            fid = f"{id_prefix}-{sheet_name[:3].upper()}-{num:03d}" if isinstance(num, int) else f"{id_prefix}-{sheet_name[:3].upper()}-{row_idx:03d}"
            apps = _app_codename_from_label(app_label) or _app_codename_from_label(title)
            severity = _normalize_severity(severidad)
            description = str(resumen).strip() if resumen else ""
            source = str(fuente).strip() if fuente else ""

            full_text = f"{title} {description}"
            is_blocker = bool(_BLOCKER_RE.search(full_text)) or severity in ("critical",)
            is_quick_win = bool(_QUICK_WIN_RE.search(full_text))
            is_correction = bool(_CORRECTION_RE.search(full_text))

            findings.append(DiscoveryFinding(
                id=fid,
                title=title[:200],
                state="CONFIRMED-BY-DOCS",
                domains=[domain_guess] if domain_guess else [],
                apps_affected=apps,
                severity=severity,
                description=description[:500],
                source_files=[f"{path.name}:{sheet_name}"] + ([source[:100]] if source else []),
                is_blocker=is_blocker,
                is_quick_win=is_quick_win,
                is_correction=is_correction,
            ))

    wb.close()
    logger.info("Parsed %d findings from %s", len(findings), path.name)
    return findings


def _domain_from_sheet_name(sheet_name: str) -> str:
    """Guess a domain tag from a spreadsheet sheet name.

    Same naming convention as ``_app_codename_from_label``: the IF-test
    strings on the left are incoming spreadsheet sheet names from
    external Excel files; the returned strings are sanitized domain
    tags, never real product/system names.
    """
    s = sheet_name.lower()
    if "sicofav" in s or "p1" in s:
        return "trazabilidad"
    if "srg" in s or "p2" in s:
        return "ventas"
    if "arc" in s or "bsp" in s or "reembolso" in s:
        return "reembolsos"
    if "cfdi" in s or "p4" in s:
        return "facturas-electronicas"
    if "comision" in s or "p5" in s:
        return "comisiones"
    if "praxis" in s or "ecosistema" in s:
        return "praxis-ecosystem"
    return ""


def load_entregables_context(entregables_path: str | Path) -> list[DiscoveryFinding]:
    """Load discovery findings from ``entregables-XXabr/`` canonical Excels.

    Looks for known workbook names:
      - ``Hallazgos_Trazabilidad_5_Apps.xlsx`` (primary — structured findings)
      - ``Panorama_Completo_5_Aplicativos.xlsx`` (secondary — context data)

    Args:
        entregables_path: directory containing the Excels.

    Returns:
        A list of ``DiscoveryFinding`` extracted from the workbooks.
        Returns an empty list if the directory or openpyxl is missing.
    """
    base = Path(entregables_path)
    if not base.is_dir():
        return []
    if not _OPENPYXL_AVAILABLE:
        logger.info("openpyxl not installed — entregables ingestion disabled")
        return []

    findings: list[DiscoveryFinding] = []

    # Primary: trazabilidad workbook
    for name in ("Hallazgos_Trazabilidad_5_Apps.xlsx",):
        p = base / name
        if p.is_file():
            findings.extend(_parse_trazabilidad_xlsx(p, id_prefix="TRZ"))

    return findings


# ── Main loader ───────────────────────────────────────────────���───────


def load_discovery_context(
    base_path: str | Path,
) -> DiscoveryIndex:
    """Load the discovery index from a directory containing ``analisis/``.

    The loader tries, in order:
      1. ``analisis/fase-3-maestro.md`` — the consolidated cross-domain
         deduped index. This is the best source when available.
      2. ``analisis/fase-2-por-dominio/*.md`` — per-domain files. Used
         as fallback when the maestro is empty or missing, or to
         supplement with domain-local findings not yet promoted.

    The maestro takes precedence: if both are present and a finding
    appears in both, only the maestro version is kept (deduplicated
    by finding ID).

    Args:
        base_path: path to the directory that contains ``analisis/``.
            Typically ``NexusForge-Shared/transcripciones-generales/``
            or a test fixture directory.

    Returns:
        A ``DiscoveryIndex`` with all parsed findings and computed
        statistics. Returns an empty index (not an error) if no
        analysis files are found.
    """
    base = Path(base_path)
    analysis_dir = base / "analisis"
    index = DiscoveryIndex(source_path=str(base))

    if not analysis_dir.is_dir():
        logger.info("No analisis/ directory found at %s", base)
        return index

    seen_ids: set[str] = set()

    # 1. Try the maestro first
    maestro_path = analysis_dir / "fase-3-maestro.md"
    if maestro_path.is_file():
        try:
            text = maestro_path.read_text(encoding="utf-8", errors="replace")
            maestro_findings = _parse_maestro(text)
            for f in maestro_findings:
                if f.id not in seen_ids:
                    index.findings.append(f)
                    seen_ids.add(f.id)
            logger.info(
                "Loaded %d findings from maestro (%s)",
                len(maestro_findings), maestro_path,
            )
        except Exception as exc:
            logger.warning("Failed to parse maestro: %s", exc)

    # 1b. Supplement with entregables/ canonical Excels if present
    #     (e.g., analisis/entregables-13abr/Hallazgos_Trazabilidad_5_Apps.xlsx)
    for entregables_dir in sorted(analysis_dir.glob("entregables-*")):
        if entregables_dir.is_dir():
            ent_findings = load_entregables_context(entregables_dir)
            added = 0
            for f in ent_findings:
                if f.id not in seen_ids:
                    index.findings.append(f)
                    seen_ids.add(f.id)
                    added += 1
            if added:
                logger.info(
                    "Added %d findings from %s (entregables)",
                    added, entregables_dir.name,
                )

    # 2. Supplement with domain files
    domain_dir = analysis_dir / "fase-2-por-dominio"
    if domain_dir.is_dir():
        for domain_file in sorted(domain_dir.glob("*.md")):
            if domain_file.name.startswith("_"):
                continue  # skip templates
            domain_name = domain_file.stem  # e.g., "ventas", "reembolsos"
            try:
                text = domain_file.read_text(encoding="utf-8", errors="replace")
                domain_findings = _parse_domain_file(text, domain_name)
                added = 0
                for f in domain_findings:
                    if f.id not in seen_ids:
                        index.findings.append(f)
                        seen_ids.add(f.id)
                        added += 1
                if added:
                    logger.debug(
                        "Added %d findings from domain %s (not in maestro)",
                        added, domain_name,
                    )
            except Exception as exc:
                logger.warning("Failed to parse domain file %s: %s", domain_file, exc)

    # 3. Compute statistics
    all_domains: set[str] = set()
    all_apps: set[str] = set()
    for f in index.findings:
        all_domains.update(f.domains)
        all_apps.update(f.apps_affected)

    index.total_findings = len(index.findings)
    index.total_validated = sum(1 for f in index.findings if f.state in ("VALIDATED", "CONFIRMED-BY-DOCS", "MAPPED"))
    index.total_blockers = sum(1 for f in index.findings if f.is_blocker)
    index.total_quick_wins = sum(1 for f in index.findings if f.is_quick_win)
    index.total_corrections = sum(1 for f in index.findings if f.is_correction)
    index.domains_covered = sorted(all_domains)
    index.apps_mentioned = sorted(all_apps)

    logger.info(
        "Discovery index loaded: %d findings (%d validated, %d blockers, "
        "%d quick wins, %d corrections) across %d domains, %d apps",
        index.total_findings, index.total_validated, index.total_blockers,
        index.total_quick_wins, index.total_corrections,
        len(index.domains_covered), len(index.apps_mentioned),
    )

    return index
